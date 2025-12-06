import tkinter as tk
from tkinter import filedialog, messagebox
import os
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl import load_workbook


class ExcelIDApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 自动填充唯一 ID 工具（最大ID+1 + 仅追加行）")
        self.root.geometry("500x300")

        # 选择文件提示
        self.label = tk.Label(root, text="请选择 Excel 文件", font=("Arial", 12))
        self.label.pack(pady=10)

        # 规则提示
        self.tip_label = tk.Label(
            root,
            text=(
                "仅对【从最后一条旧数据之后追加的新行】生成 ID\n"
                "插入在中间的行不会生成 ID\n"
                "满足 C+D 或 A+B+C+D 任一才写入 ID\n"
                "导入表格可能会卡顿，请耐心等待"
            ),
            font=("Arial", 10),
            justify="left",
            wraplength=480,
        )
        self.tip_label.pack(pady=5)

        # 拖拽
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind("<<Drop>>", self.handle_drop)

        self.select_button = tk.Button(root, text="选择文件", command=self.load_file, width=20)
        self.select_button.pack()

        self.process_button = tk.Button(
            root,
            text="生成 ID 并导出",
            command=self.process_file,
            width=20,
            state=tk.DISABLED,
        )
        self.process_button.pack(pady=15)

        self.file_path = None

    def handle_drop(self, event):
        filepath = event.data.strip("{}")
        if os.path.isfile(filepath):
            self.file_path = filepath
            self.label.config(text=f"已拖入：{os.path.basename(filepath)}")
            self.process_button.config(state=tk.NORMAL)
        else:
            messagebox.showerror("错误", "拖拽的不是文件")

    def load_file(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filepath:
            self.file_path = filepath
            self.label.config(text=f"已选择：{os.path.basename(filepath)}")
            self.process_button.config(state=tk.NORMAL)

    def process_file(self):
        if not self.file_path:
            messagebox.showerror("错误", "请先选择文件")
            return

        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            # ============ 1. 找到已有 ID 的最大 ID 和 最后 ID 所在行 ============
            existing_ids = []
            last_id_row = 1  # 记录“最后一个 ID 的行号”

            for i in range(2, ws.max_row + 1):
                j = ws.cell(i, 10).value
                if j and str(j).isdigit():
                    existing_ids.append(int(str(j)))
                    last_id_row = i

            # ============ >>> 新增代码：检测是否有重复 ID ============
            id_set = set()
            duplicate_ids = set()

            for id_val in existing_ids:
                if id_val in id_set:
                    duplicate_ids.add(id_val)
                else:
                    id_set.add(id_val)

            if duplicate_ids:
                messagebox.showerror(
                    "检测到重复 ID",
                    f"Excel 中出现重复 ID：{sorted(list(duplicate_ids))}\n\n"
                    "请修复后再导出，否则导入数据库会失败。"
                )
                return  # 直接终止
            # ============ <<< 新增代码结束 ============

            max_id = max(existing_ids) if existing_ids else 0
            next_id = max_id + 1

            # ============ 2. 从 last_id_row+1 开始才允许加 ID ============
            for i in range(last_id_row + 1, ws.max_row + 1):
                a = ws.cell(i, 1).value
                b = ws.cell(i, 2).value
                c = ws.cell(i, 3).value
                d = ws.cell(i, 4).value
                jcell = ws.cell(i, 10)

                trigger = (c and d) or (a and b and c and d)

                # 必须满足条件 AND J 列为空
                if trigger and (jcell.value is None or str(jcell.value).strip() == ""):
                    jcell.value = next_id
                    next_id += 1

            # 保存文件
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
            )

            if save_path:
                wb.save(save_path)
                messagebox.showinfo("成功", "导出完成！")

        except Exception as e:
            messagebox.showerror("错误", f"处理失败：\n{str(e)}")


if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = ExcelIDApp(root)
    root.mainloop()